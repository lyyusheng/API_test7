# coding:utf-8
from openpyxl import load_workbook

from API_7.common import project_path
from API_7.common.read_config import ReadConfig


class DoExcel:
    """该类用于读取excel中的测试数据，以及测试数据的写回"""

    def __init__(self, file_name, sheet_name):
        """file_name:工作簿名字或者路径,sheet_name:表单名"""
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_case(self, section):
        """此函数用于读取Excel里面的测试用例，有返回值
        section: 配置文件里面的片段名，根据设定来具体执行哪些测试用例"""

        # 配置文件控制读取哪些用例
        case_id = ReadConfig(project_path.conf_path).get_data(section,
                                                              'case_id')  # get_data()是自己写的read_config模块中用于读取配置文件中元组、字典、列表的函数
        wb = load_workbook(self.file_name)  # 打开工作簿,文件名出现了t开头，要多加/
        sheet = wb[self.sheet_name]  # 定位表单

        # tel = self.get_tel()  # 放这里不妥，要是多条用例需要新的手机号，却还没更新
        test_data = []
        for i in range(2, sheet.max_row + 1):
            # tel = self.get_tel()  #放这里也不妥，不需要新手机号也会每次都更新，浪费！
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Method'] = sheet.cell(i, 4).value
            row_data['Url'] = sheet.cell(i, 5).value
            # 自动更新手机号
            if sheet.cell(i, 6).value.find('tel') != -1:  # find()是内置函数，如果找到'tel'则返回最小索引值，找不到则返回-1.也可以用成员运算in
                tel = self.get_tel()  # 对象调用get_tel()方法获取存在excel里面的电话号码
                row_data['Params'] = sheet.cell(i, 6).value.replace('tel', str(
                    tel))  # replace是字符串之间的替换，tel是int，所以用str(tel)转str
                self.update_tel(int(tel) + 1)  # 将手机号+1再写回excel中的备用手机号 要是同时有两条用例需要没注册过的手机号？
            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()  # 记得关闭

        # 根据配置文件读取测试用例
        final_data = []  # 定义一个空列表，用于存储根据配置文件需要读取的测试用例
        if case_id == "all":  # 把1换成all时报错：不可迭代,配置文件把all变成"all"即可，本来是str,都出来也是str，相当于双重str
            final_data = test_data
        else:
            for i in case_id:  # case_id是配置文件[CASE]里面的section,具体值可自行配置
                final_data.append(test_data[i - 1])  # i是case，从1开始，但是列表索引是从0开始遍历的，所以要减1,但是如果配置文件中的i是"all",
        return final_data

    def get_tel(self):
        """读取excel里面的手机号码"""
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()  # 记得关闭
        return sheet.cell(1, 2).value

    def update_tel(self, new_tel):
        """用于更新excel里面的备用电话号码"""
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        sheet.cell(1, 2, new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self, row, col, value):
        """写回测试结果"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    de = DoExcel(project_path.case_path, 'add_loan')  # 注意有时要多加/
    # res = de.read_case('AddLoanCASE')
    # print(res)
    # de.write_back(2,8,'试一下能不能写回')#执行前需要手动关闭Excel，否则报错
