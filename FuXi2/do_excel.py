#1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
# 函数一：读取指定表单的数据，
# 有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。 每一行都有 一
# 个单独的row_list [[1,2,3],[4,5,6]] #每一行数据读取完毕后，把row_list存到大列表all_row_list
#函数二：在指定的单元格写入指定的数据，并保存到当前Excel
#函数三：新建一个Excel
from openpyxl import load_workbook  #load_workbook用于打开工作簿
from openpyxl import  workbook
from configparser import ConfigParser
from FuXi2.read_config import ReadConfig
from FuXi2.my_log import MyLog

logger=MyLog()
# 创建对象：
cf = ConfigParser()
# 打开文件 read（）
cf.read('lemon.conf', encoding='utf-8')
button = cf.get('TestCase', 'button')

class DoExcel():
    '''类的作用是完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name):#初始化函数,实现参数化.后面函数不用再传
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_case(self):
        '''读取指定表单的数据，以嵌套的形式存储'''
        wb=load_workbook(self.file_name)#定位工作簿,初始函数参数化，用到初始化的变量前面都要加self.
        sheet=wb[self.sheet_name]        #定位表单
    #嵌套列表
        #row_list=[]    放在这里是错的，每存完一行数据，并且存到大列表之后，row_list这个列表都要被重置为空列表
        logger.info('开始读数据啦！')
        test_data=[]
        if button=='1': #1读取所有用例，注意：配置文件那里不要写成[1],是1,看清楚配置文件写的是啥
            for i in range(2,sheet.max_row+1):
                row_list = []   #每存完一行数据，并且存到大列表之后，row_list这个列表都要被重置为空列表
                for j in range(1,sheet.max_column+1):
                    res=sheet.cell(row=i,column=j).value
                    row_list.append(res)
                test_data.append(row_list)
            logger.info('数据读取完毕！')
        else:
            for i in eval(button):#eval()之后变成一个列表
                row_list = []  # 每存完一行数据，并且存到大列表之后，row_list这个列表都要被重置为空列表
                for j in range(1, sheet.max_column + 1):
                    res = sheet.cell(row=i+1, column=j).value
                    row_list.append(res)
                test_data.append(row_list)
            logger.info('数据读取完毕！')
        return test_data
    #嵌套字典
        # test_data=[]
        # for i in range(2,sheet.max_row+1):
        #     row_list={}
        #     row_list['CaseId']=sheet.cell(i,1).value
        #     row_list['Title'] = sheet.cell(i, 2).value
        #     row_list['Module'] = sheet.cell(i,3).value
        #     row_list['TestData'] = sheet.cell(i,4).value
        #     row_list['ExpectedResult'] = sheet.cell(i,5).value
        #     row_list['ActualResult'] = sheet.cell(i,6).value
        #     row_list['TestResult'] = sheet.cell(i,7).value
        #     test_data.append(row_list)
        # return test_data

    def write_case(self,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)#打开要写入的数据的工作簿
        sheet=wb[self.sheet_name]    #定位表单
        logger.info('开始往excel写入数据')
        try:
            sheet.cell(row,col).value=value #定位单元格并往单元格写入数据，或者第二种方法sheet.cell(row,col,value)
            wb.save(self.file_name)  #保存工作簿
            wb.close()  #每次操作完必须关闭掉
        except:
            logger.error('往excel写入数据完毕')
        logger.info('往excel写入数据完毕')
    def create_wb(self):
        '''新建一个Excel'''
        pass
if __name__ == '__main__':

    de=DoExcel('py14excel.xlsx','test_case')#写了初始化函数，被初始化的参数都从对象传入，而不是在下面调用的函数
    test_datas=de.read_case()#被初始化的参数不再从调用的函数传入，从对象传入
    print(test_datas)
