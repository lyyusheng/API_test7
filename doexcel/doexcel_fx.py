# from openpyxl import load_workbook  #load_workbook用于打开工作簿,支持读、写操作
# #wb1=load_workbook('E:\Python14\Homework\python_hw.xlsx')#绝对路径,开头遇到转义字符要加多一个‘/’，例如/20，要变成//20
# wb=load_workbook('python_hw.xlsx')
# #定位表单
# #sheet_1=wb.get_sheet_by_name('test_sheet')#旧的方法，已经弃用
# sheet=wb['test_sheet']  #常用这个
#
# #定位单元格
# for j in range(1,sheet.max_row+1):
#     for i in range(1,sheet.max_column+1):
#         res=sheet.cell(row=j,column=i).value#或者cell（2，3）不指定则默认为cell（row=,column=）
#         if res!=None:#排除None
#             print(res)
#写数据
#方法一：
#sheet.cell(3,2,'写操作的内容：写个大美女')
#方法二：
# sheet.cell(6,2).value='第二种写操作：似乎更好用'
#
# wb.save('python_hw.xlsx')   #文件名不存在则另存为，如果是原有的文件名则要手动关闭Excel
#必须手动关闭Excel，否则报错，必须手动！！！
#新建工作簿
from openpyxl import workbook
wb=workbook.Workbook()
wb.save('py14excel.xlsx')
#新建的要重新打開才能操作